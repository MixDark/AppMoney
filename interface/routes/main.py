from flask import render_template, request, redirect, url_for, session, flash, send_file, jsonify, current_app
from flask import Blueprint
from functools import wraps
import base64
import os
from infrastructure.db import get_connection
from infrastructure.validators import *
from application.currencies import get_currency_list, get_default_value, get_currency_display, get_currency_info
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date
import calendar
import secrets
from io import BytesIO
import time
from threading import Lock
import pandas as pd
import traceback
from captcha.image import ImageCaptcha
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import pyotp
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def login_required_custom(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('main.login'))
        return f(*args, **kwargs)
    return decorated_function

def detect_image_type(data):
    if data.startswith(b'\xff\xd8\xff'):
        return 'jpeg'
    if data.startswith(b'\x89PNG\r\n\x1a\n'):
        return 'png'
    if data.startswith(b'GIF87a') or data.startswith(b'GIF89a'):
        return 'gif'
    if data.startswith(b'RIFF') and data[8:12] == b'WEBP':
        return 'webp'
    return None

def get_csrf_token():
    token = session.get('csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['csrf_token'] = token
    return token

def generar_codigo_captcha(length=5):
    caracteres = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return ''.join(secrets.choice(caracteres) for _ in range(length))

def generar_captcha_imagen(codigo):
    captcha = ImageCaptcha(width=320, height=120)
    buffer = BytesIO()
    captcha.write(codigo, buffer)
    imagen = buffer.getvalue()
    return 'data:image/png;base64,' + base64.b64encode(imagen).decode('ascii')

def crear_captcha_challenge(token):
    captcha_code = generar_codigo_captcha(6)
    session[f'captcha_{token}'] = captcha_code
    return {
        'captcha_code': captcha_code,
        'captcha_image': generar_captcha_imagen(captcha_code),
    }

_rate_limit_lock = Lock()
_rate_limit_state = {}

def rate_limited(key, limit, window_seconds):
    now = time.time()
    with _rate_limit_lock:
        timestamps = _rate_limit_state.get(key, [])
        timestamps = [ts for ts in timestamps if now - ts < window_seconds]
        if len(timestamps) >= limit:
            _rate_limit_state[key] = timestamps
            return True
        timestamps.append(now)
        _rate_limit_state[key] = timestamps
        return False

def read_profile_image(file):
    if not file or not file.filename:
        return None, None
    max_bytes = current_app.config.get('PHOTO_MAX_BYTES', 2 * 1024 * 1024)
    data = file.read(max_bytes + 1)
    if len(data) > max_bytes:
        return None, "La imagen supera el tamaño permitido"
    kind = detect_image_type(data)
    if kind not in {'jpeg', 'png', 'gif', 'webp'}:
        return None, "Formato de imagen no permitido"
    return data, None

def _serialize_usuario_cache(usuario):
    if not usuario:
        return None

    cached_usuario = {
        'id': usuario.get('id'),
        'nombre_usuario': usuario.get('nombre_usuario'),
        'nombre_completo': usuario.get('nombre_completo'),
        'email': usuario.get('email'),
        'telefono': usuario.get('telefono'),
        'pais': usuario.get('pais'),
        'ciudad': usuario.get('ciudad'),
        'moneda': usuario.get('moneda', 'USD'),
        'MFA': bool(usuario.get('MFA')),
        'ultimo_login': usuario.get('ultimo_login'),
        'foto_perfil': bool(usuario.get('foto_perfil')),
        'creado_en': usuario.get('creado_en'),
    }
    return cached_usuario


def _store_usuario_cache(usuario):
    cached_usuario = _serialize_usuario_cache(usuario)
    if cached_usuario is not None:
        session['usuario_cache'] = cached_usuario
    else:
        session.pop('usuario_cache', None)


def get_current_usuario(refresh=False):
    if 'usuario_id' not in session:
        return None

    if not refresh:
        cached_usuario = session.get('usuario_cache')
        if cached_usuario and cached_usuario.get('id') == session.get('usuario_id'):
            return cached_usuario

    usuario = obtener_usuario_por_id(session['usuario_id'])
    if usuario:
        _store_usuario_cache(usuario)
    return usuario


def _agrupar_montos_por_mes(registros):
    montos_por_mes = {}

    for registro in registros or []:
        fecha_registro = registro.get('fecha')
        if isinstance(fecha_registro, datetime):
            fecha_dt = fecha_registro
        elif isinstance(fecha_registro, date):
            fecha_dt = datetime.combine(fecha_registro, datetime.min.time())
        elif isinstance(fecha_registro, str):
            try:
                fecha_dt = datetime.fromisoformat(fecha_registro)
            except ValueError:
                try:
                    fecha_dt = datetime.strptime(fecha_registro, '%Y-%m-%d')
                except ValueError:
                    continue
        else:
            continue

        clave_mes = (fecha_dt.year, fecha_dt.month)
        montos_por_mes[clave_mes] = montos_por_mes.get(clave_mes, 0.0) + float(registro.get('monto') or 0)

    return montos_por_mes

def register_routes(app):
    bp = Blueprint('main', __name__)

    @bp.before_app_request
    def enforce_csrf():
        if request.method in {'POST', 'PUT', 'DELETE', 'PATCH'}:
            if request.endpoint and request.endpoint.startswith('static'):
                return None
            csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
            if not csrf_token or csrf_token != session.get('csrf_token'):
                return 'CSRF token inválido', 400

    @bp.app_context_processor
    def inject_csrf_token():
        context = {
            'csrf_token': get_csrf_token(),
            'currencies_list': get_currency_list()
        }
        usuario = get_current_usuario()
        if usuario:
            context['usuario'] = usuario
            context['currency_code'] = usuario.get('moneda', 'USD')
        return context

    @bp.route('/', methods=['GET'])
    def consolidado():
        if 'usuario_id' not in session:
            return redirect(url_for('main.login'))
        usuario = get_current_usuario()
        if usuario is None:
            session.clear()
            return redirect(url_for('main.login'))
        
        ingresos = obtener_ingresos(usuario['id'])
        gastos = obtener_gastos(usuario['id'])
        inversiones = obtener_inversiones(usuario['id'])
        
        total_ingresos = float(sum(i['monto'] for i in ingresos) or 0)
        total_gastos = float(sum(g['monto'] for g in gastos) or 0)
        total_inversiones = float(sum(inv['monto'] for inv in inversiones) or 0)
        saldo = total_ingresos - total_gastos - total_inversiones
        
        # Procesar transacciones recurrentes pendientes
        procesar_recurrentes_usuario(usuario['id'])
        
        # Calcular tendencias de los últimos 6 meses
        hoy = datetime.now()
        ingresos_por_mes = _agrupar_montos_por_mes(ingresos)
        gastos_por_mes = _agrupar_montos_por_mes(gastos)
        inversiones_por_mes = _agrupar_montos_por_mes(inversiones)
        meses_labels = []
        data_ingresos = []
        data_gastos = []
        data_inversiones = []
        
        for i in range(5, -1, -1):
            fecha_temp = hoy - timedelta(days=i*30)
            anio, mes = fecha_temp.year, fecha_temp.month
            meses_labels.append(fecha_temp.strftime('%b %Y'))
            
            ing_mes = ingresos_por_mes.get((anio, mes), 0.0)
            gas_mes = gastos_por_mes.get((anio, mes), 0.0)
            inv_mes = inversiones_por_mes.get((anio, mes), 0.0)
            
            data_ingresos.append(float(ing_mes))
            data_gastos.append(float(gas_mes))
            data_inversiones.append(float(inv_mes))

        # Comparación con promedios (Mock data for analysis)
        promedio_nacional_gasto = 1500.00
        total_movimientos = total_ingresos + total_gastos + total_inversiones
        
        analysis = {
            'meses_labels': meses_labels,
            'data_ingresos': data_ingresos,
            'data_gastos': data_gastos,
            'data_inversiones': data_inversiones,
            'promedio_nacional_gasto': promedio_nacional_gasto,
            'ahorro_porcentaje': round((saldo / total_ingresos * 100), 1) if total_ingresos > 0 else 0,
            'ingreso_porcentaje': round((total_ingresos / total_movimientos * 100), 1) if total_movimientos > 0 else 0,
            'gasto_porcentaje': round((total_gastos / total_movimientos * 100), 1) if total_movimientos > 0 else 0,
            'inversion_porcentaje': round((total_inversiones / total_movimientos * 100), 1) if total_movimientos > 0 else 0
        }

        return render_template('dashboard/consolidado.html', 
                             usuario=usuario, 
                             ingresos=ingresos[:10], # Limitar para el dashboard
                             gastos=gastos[:10], 
                             inversiones=inversiones[:10],
                             total_ingresos=total_ingresos, 
                             total_gastos=total_gastos, 
                             total_inversiones=total_inversiones, 
                             saldo=saldo,
                             analysis=analysis)

    @bp.route('/login', methods=['GET', 'POST'])
    def login():
        mostrar_otp_modal = False
        
        if request.method == 'POST':
            ip = request.remote_addr or 'unknown'
            if rate_limited(f'login:{ip}', limit=5, window_seconds=60):
                flash('Demasiados intentos. Intenta de nuevo más tarde', 'error')
                return render_template('auth/login.html', mostrar_otp_modal=False)
            nombre_usuario = request.form.get('nombre_usuario', '').strip()
            password = request.form.get('password', '')
            
            valid, msg = validate_username(nombre_usuario)
            if not valid:
                flash(msg, 'error')
                return render_template('auth/login.html', mostrar_otp_modal=False)
            
            usuario = obtener_usuario_por_nombre(nombre_usuario)
            
            if not usuario:
                flash('Usuario no encontrado', 'error')
            elif not check_password_hash(usuario['password_hash'], password):
                flash('Contraseña incorrecta', 'error')
            else:
                # Usuario y contraseña son correctos
                # Verificar si tiene MFA habilitado
                if usuario_tiene_mfa_activado(usuario['id']):
                    # Guardar el usuario temporalmente en sesión
                    session['temp_usuario_id'] = usuario['id']
                    session['username_temp'] = usuario['nombre_usuario']
                    mostrar_otp_modal = True
                    return render_template('auth/login.html', 
                                         mostrar_otp_modal=mostrar_otp_modal,
                                         nombre_usuario=nombre_usuario)
                else:
                    # Sin MFA, hacer login directo
                    actualizar_ultimo_login(usuario['id'])
                    session.permanent = True
                    session['usuario_id'] = usuario['id']
                    _store_usuario_cache(usuario)
                    session.pop('temp_usuario_id', None)
                    session.pop('username_temp', None)
                    return redirect(url_for('main.consolidado'))
        
        return render_template('auth/login.html', mostrar_otp_modal=False)

    @bp.route('/logout')
    def logout():
        session.clear()
        return redirect(url_for('main.login'))

    @bp.route('/usuario/<int:usuario_id>/foto')
    @login_required_custom
    def foto_perfil(usuario_id):
        if session.get('usuario_id') != usuario_id:
            return '', 403
        conn = get_connection()
        if conn is None:
            return '', 404
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT foto_perfil FROM usuarios WHERE id = %s', (usuario_id,))
            result = cursor.fetchone()
            cursor.close()
            if not result or not result[0]:
                return '', 404
            foto_bytes = result[0]
            kind = detect_image_type(foto_bytes)
            if kind == 'jpeg':
                mime = 'image/jpeg'
            elif kind:
                mime = f"image/{kind}"
            else:
                mime = 'application/octet-stream'
            response = send_file(BytesIO(foto_bytes), mimetype=mime)
            response.headers['Cache-Control'] = 'no-store'
            return response
        finally:
            conn.close()

    @bp.route('/verify-otp', methods=['POST'])
    def verify_otp():
        """Verifica el código OTP proporcionado por el usuario"""
        if 'temp_usuario_id' not in session:
            return jsonify({'exito': False, 'mensaje': 'Sesión inválida'}), 400

        ip = request.remote_addr or 'unknown'
        if rate_limited(f'otp:{ip}', limit=5, window_seconds=60):
            return jsonify({'exito': False, 'mensaje': 'Demasiados intentos. Intenta de nuevo más tarde'}), 429
        
        codigo_otp = request.form.get('codigo_otp', '').strip()
        usuario_id = session['temp_usuario_id']
        
        valid, msg = validate_totp(codigo_otp)
        if not valid:
            return jsonify({'exito': False, 'mensaje': msg}), 400
        
        # Obtener el secreto TOTP del usuario
        totp_secret = obtener_totp_secret_usuario(usuario_id)
        
        if not totp_secret:
            return jsonify({'exito': False, 'mensaje': 'Error: MFA no configurado correctamente'}), 400
        
        # Validar el código OTP
        if validar_totp(totp_secret, codigo_otp):
            # OTP correcto, completar el login
            usuario = obtener_usuario_por_id(usuario_id)
            if usuario:
                actualizar_ultimo_login(usuario_id)
                session.permanent = True
                session['usuario_id'] = usuario_id
                _store_usuario_cache(usuario)
                session.pop('temp_usuario_id', None)
                session.pop('username_temp', None)
                return jsonify({'exito': True, 'mensaje': 'OTP verificado correctamente', 'redirect': url_for('main.consolidado')})
            else:
                return jsonify({'exito': False, 'mensaje': 'Error: Usuario no encontrado'}), 400
        else:
            # OTP incorrecto
            return jsonify({'exito': False, 'mensaje': 'Código OTP inválido'}), 401

    @bp.route('/recuperar_contraseña', methods=['GET', 'POST'])
    def recuperar_contraseña():
        exito = None
        if request.method == 'POST':
            ip = request.remote_addr or 'unknown'
            if rate_limited(f'recovery:{ip}', limit=3, window_seconds=300):
                flash('Demasiados intentos. Intenta de nuevo más tarde', 'error')
                return render_template('auth/recuperar_contraseña.html')
            nombre_usuario = request.form.get('nombre_usuario', '').strip()
            
            valid, msg = validate_username(nombre_usuario)
            if not valid:
                flash(msg, 'error')
                return render_template('auth/recuperar_contraseña.html')
            
            usuario = obtener_usuario_por_nombre(nombre_usuario)
            
            if not usuario:
                flash('Usuario no encontrado', 'error')
            else:
                # Generar token único
                token = secrets.token_urlsafe(32)
                expires_at = datetime.now() + timedelta(hours=1)
                
                # Guardar token en BD
                if crear_token_recuperacion(usuario['id'], token, expires_at):
                    # Redirigir al CAPTCHA
                    return redirect(url_for('main.verificar_captcha', token=token))
                else:
                    flash('Error al procesar tu solicitud', 'error')
        
        return render_template('auth/recuperar_contraseña.html', exito=exito)

    @bp.route('/verificar_captcha', methods=['GET', 'POST'])
    def verificar_captcha():
        token = request.args.get('token') or request.form.get('token')
        
        if not token:
            flash('Token inválido', 'error')
            return redirect(url_for('main.recuperar_contraseña'))
        
        # Verificar que el token sea válido
        token_data = obtener_token_recuperacion(token)
        
        if not token_data:
            flash('Token inválido o expirado', 'error')
            return redirect(url_for('main.recuperar_contraseña'))
        
        # Verificar si el token ha expirado
        expires_at = token_data['expires_at']
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        
        if expires_at < datetime.now():
            flash('El enlace de recuperación ha expirado', 'error')
            eliminar_token_recuperacion(token)
            return redirect(url_for('main.recuperar_contraseña'))
        
        # Si el token ya fue verificado, redirigir al reseteo
        if token_data['verified']:
            return redirect(url_for('main.resetear_contraseña', token=token))
        
        if request.method == 'GET':
            captcha = crear_captcha_challenge(token)

            return render_template('auth/verificar_captcha.html',
                                 token=token,
                                 captcha_image=captcha['captcha_image'])
        
        elif request.method == 'POST':
            ip = request.remote_addr or 'unknown'
            if rate_limited(f'captcha:{ip}', limit=5, window_seconds=300):
                flash('Demasiados intentos. Intenta de nuevo más tarde', 'error')
                captcha = crear_captcha_challenge(token)
                return render_template('auth/verificar_captcha.html',
                                     token=token,
                                     captcha_image=captcha['captcha_image'])
            respuesta_usuario = request.form.get('captcha_answer', '').strip().upper()
            respuesta_correcta = session.get(f'captcha_{token}')
            
            if respuesta_usuario == respuesta_correcta:
                # CAPTCHA correcto, marcar token como verificado
                actualizar_token_verificado(token)
                session.pop(f'captcha_{token}', None)
                if usuario_tiene_mfa_activado(token_data['usuario_id']):
                    session.pop(f'recovery_otp_{token}', None)
                    return redirect(url_for('main.verificar_otp_recuperacion', token=token))

                session.pop(f'recovery_otp_{token}', None)
                return redirect(url_for('main.resetear_contraseña', token=token))
            else:
                flash('Respuesta incorrecta. Intenta de nuevo', 'error')
                
                # Generar nuevo CAPTCHA
                captcha = crear_captcha_challenge(token)

                return render_template('auth/verificar_captcha.html',
                                     token=token,
                                     captcha_image=captcha['captcha_image'])

    @bp.route('/verificar_captcha/nuevo', methods=['GET'])
    def nuevo_captcha():
        token = request.args.get('token')

        if not token:
            return jsonify({'ok': False, 'error': 'Token inválido'}), 400

        token_data = obtener_token_recuperacion(token)
        if not token_data:
            return jsonify({'ok': False, 'error': 'Token inválido o expirado'}), 400

        captcha = crear_captcha_challenge(token)
        return jsonify({'ok': True, 'captcha_image': captcha['captcha_image']})

    @bp.route('/verificar_otp_recuperacion', methods=['GET', 'POST'])
    def verificar_otp_recuperacion():
        token = request.args.get('token') or request.form.get('token')

        if not token:
            flash('Token inválido', 'error')
            return redirect(url_for('main.recuperar_contraseña'))

        token_data = obtener_token_recuperacion(token)
        if not token_data:
            flash('Token inválido o expirado', 'error')
            return redirect(url_for('main.recuperar_contraseña'))

        expires_at = token_data['expires_at']
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)

        if expires_at < datetime.now():
            flash('El enlace de recuperación ha expirado', 'error')
            eliminar_token_recuperacion(token)
            return redirect(url_for('main.recuperar_contraseña'))

        if not token_data['verified']:
            return redirect(url_for('main.verificar_captcha', token=token))

        if not usuario_tiene_mfa_activado(token_data['usuario_id']):
            session.pop(f'recovery_otp_{token}', None)
            return redirect(url_for('main.resetear_contraseña', token=token))

        if request.method == 'POST':
            ip = request.remote_addr or 'unknown'
            if rate_limited(f'recovery_otp:{ip}', limit=5, window_seconds=300):
                flash('Demasiados intentos. Intenta de nuevo más tarde', 'error')
                return render_template('auth/verificar_otp_recuperacion.html', token=token)

            codigo_otp = request.form.get('codigo_otp', '').strip()
            valid, msg = validate_totp(codigo_otp)
            if not valid:
                flash(msg, 'error')
                return render_template('auth/verificar_otp_recuperacion.html', token=token)

            totp_secret = obtener_totp_secret_usuario(token_data['usuario_id'])
            if not totp_secret:
                flash('Error: MFA no configurado correctamente', 'error')
                return render_template('auth/verificar_otp_recuperacion.html', token=token)

            if validar_totp(totp_secret, codigo_otp):
                session[f'recovery_otp_{token}'] = True
                return redirect(url_for('main.resetear_contraseña', token=token))

            flash('Código OTP inválido', 'error')
            return render_template('auth/verificar_otp_recuperacion.html', token=token)

        return render_template('auth/verificar_otp_recuperacion.html', token=token)

    @bp.route('/resetear_contraseña', methods=['GET', 'POST'])
    def resetear_contraseña():
        token = request.args.get('token') or request.form.get('token')
        
        if not token:
            flash('Token inválido', 'error')
            return render_template('auth/resetear_contraseña.html', token='')
        
        # Verificar que el token sea válido
        token_data = obtener_token_recuperacion(token)
        
        if not token_data:
            flash('Token inválido o expirado', 'error')
            return render_template('auth/resetear_contraseña.html', token='')
        
        # Verificar si el token ha expirado
        expires_at = token_data['expires_at']
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        
        if expires_at < datetime.now():
            flash('El enlace de recuperación ha expirado', 'error')
            eliminar_token_recuperacion(token)
            return render_template('auth/resetear_contraseña.html', token='')
        
        # Verificar que el CAPTCHA fue verificado
        if not token_data['verified']:
            return redirect(url_for('main.verificar_captcha', token=token))

        if usuario_tiene_mfa_activado(token_data['usuario_id']) and not session.get(f'recovery_otp_{token}'):
            return redirect(url_for('main.verificar_otp_recuperacion', token=token))
        
        if request.method == 'POST':
            ip = request.remote_addr or 'unknown'
            if rate_limited(f'reset:{ip}', limit=5, window_seconds=300):
                flash('Demasiados intentos. Intenta de nuevo más tarde', 'error')
                return render_template('auth/resetear_contraseña.html', token=token)
            password_nueva = request.form.get('password_nueva', '')
            password_confirmar = request.form.get('password_confirmar', '')
            
            valid, msg = validate_password(password_nueva)
            if not valid:
                flash(msg, 'error')
                return render_template('auth/resetear_contraseña.html', token=token)
            
            if password_nueva != password_confirmar:
                flash('Las contraseñas no coinciden', 'error')
                return render_template('auth/resetear_contraseña.html', token=token)
            
            # Actualizar contraseña
            nueva_password_hash = generate_password_hash(password_nueva)
            if actualizar_password_usuario(token_data['usuario_id'], nueva_password_hash):
                # Eliminar token usado
                eliminar_token_recuperacion(token)
                session.pop(f'recovery_otp_{token}', None)
                flash('Contraseña actualizada exitosamente. Inicia sesión con tu nueva contraseña.', 'success')
                return redirect(url_for('main.login'))
            else:
                flash('Error al actualizar la contraseña', 'error')
        
        return render_template('auth/resetear_contraseña.html', token=token)

    @bp.route('/registrar', methods=['GET', 'POST'])
    @login_required_custom
    def registrar():
        if request.method == 'POST':
            tipo = request.form.get('tipo', '').strip()
            monto_str = request.form.get('monto', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            fecha = request.form.get('fecha', '').strip()
            usuario_id = session.get('usuario_id')
            
            valid, msg = validate_tipo(tipo)
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.registrar'))
            
            valid, msg = validate_monto(monto_str)
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.registrar'))
            
            valid, msg = validate_descripcion(descripcion)
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.registrar'))
            
            valid, msg = validate_fecha(fecha)
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.registrar'))
            
            monto = float(monto_str)
            categoria_id = request.form.get('categoria_id') or None
            if categoria_id:
                try:
                    categoria_id = int(categoria_id)
                except ValueError:
                    categoria_id = None

            if tipo == 'ingreso':
                agregar_ingreso(usuario_id, monto, descripcion, fecha, categoria_id)
            elif tipo == 'gasto':
                agregar_gasto(usuario_id, monto, descripcion, fecha, categoria_id)
            elif tipo == 'inversion':
                agregar_inversion(usuario_id, '', monto, descripcion, fecha, categoria_id)
            
            flash(f'{tipo.capitalize()} registrado correctamente', 'success')
            return redirect(url_for('main.consolidado'))
        
        usuario_id = session.get('usuario_id')
        usuario = obtener_usuario_por_id(usuario_id)
        
        ingresos = obtener_ingresos(usuario_id)
        gastos = obtener_gastos(usuario_id)
        inversiones = obtener_inversiones(usuario_id)
        categorias = obtener_categorias(usuario_id)
        
        # Obtener el valor predeterminado para la moneda del usuario
        default_value = get_default_value(usuario['moneda']) if usuario else 30
        
        return render_template('forms/registrar.html', 
                             usuario=usuario,
                             ingresos=ingresos, 
                             gastos=gastos, 
                             inversiones=inversiones, 
                             categorias=categorias,
                             default_value=default_value)

    @bp.route('/reportes', methods=['GET'])
    def reportes():
        if 'usuario_id' not in session:
            return redirect(url_for('main.login'))
        usuario_id = session['usuario_id']
        mes = request.args.get('mes')
        if mes:
            anio, mes_num = mes.split('-')
            ingresos = obtener_ingresos_por_mes(usuario_id, anio, mes_num)
            gastos = obtener_gastos_por_mes(usuario_id, anio, mes_num)
            inversiones = obtener_inversiones_por_mes(usuario_id, anio, mes_num)
            desde_mes, hasta_mes = calcular_rango_mes(anio, mes_num)
            mes_mostrar = f'{_formatear_fecha_ddmmaa(desde_mes)} - {_formatear_fecha_ddmmaa(hasta_mes)}'
        else:
            hoy = datetime.now()
            anio, mes_num = hoy.year, hoy.month
            ingresos = obtener_ingresos_por_mes(usuario_id, anio, mes_num)
            gastos = obtener_gastos_por_mes(usuario_id, anio, mes_num)
            inversiones = obtener_inversiones_por_mes(usuario_id, anio, mes_num)
            desde_mes, hasta_mes = calcular_rango_mes(anio, mes_num)
            mes_mostrar = f'{_formatear_fecha_ddmmaa(desde_mes)} - {_formatear_fecha_ddmmaa(hasta_mes)}'
        total_ingresos = float(sum(i['monto'] for i in ingresos) or 0)
        total_gastos = float(sum(g['monto'] for g in gastos) or 0)
        total_inversiones = float(sum(inv['monto'] for inv in inversiones) or 0)
        saldo = total_ingresos - total_gastos - total_inversiones
        return render_template('reports/reportes.html', ingresos=ingresos, gastos=gastos, inversiones=inversiones, total_ingresos=total_ingresos, total_gastos=total_gastos, total_inversiones=total_inversiones, saldo=saldo, mes_actual=f'{anio}-{str(mes_num).zfill(2)}', mes_mostrar=mes_mostrar)

    @bp.route('/inversiones', methods=['GET', 'POST'])
    def inversiones():
        if 'usuario_id' not in session:
            return redirect(url_for('main.login'))
        if request.method == 'POST':
            tipo = request.form.get('tipo', 'otro').strip()
            descripcion = request.form.get('descripcion', '').strip()
            fecha = request.form.get('fecha', '').strip()
            usuario_id = session.get('usuario_id')
            if not usuario_id:
                return redirect(url_for('main.login'))
            valid, msg = validate_monto(request.form.get('monto', ''))
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.inversiones'))
            valid, msg = validate_fecha(fecha)
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.inversiones'))
            monto = float(request.form.get('monto'))
            categoria_id = request.form.get('categoria_id') or None
            if categoria_id:
                try:
                    categoria_id = int(categoria_id)
                except ValueError:
                    categoria_id = None

            if agregar_inversion(usuario_id, tipo, monto, descripcion, fecha, categoria_id):
                flash('Inversión registrada correctamente', 'success')
            else:
                flash('No se pudo registrar la inversión. Revisa la conexión a la base de datos.', 'error')
            return redirect(url_for('main.inversiones'))
        usuario_id = session['usuario_id']
        inversiones_list = obtener_inversiones(usuario_id)
        total_inversiones = float(sum(i['monto'] for i in inversiones_list) or 0)
        categorias = obtener_categorias(usuario_id)
        return render_template('dashboard/inversiones.html', inversiones=inversiones_list, total_inversiones=total_inversiones, categorias=categorias)

    @bp.route('/registrar_usuario', methods=['GET', 'POST'])
    def registrar_usuario():
        if request.method == 'POST':
            try:
                nombre_usuario = request.form.get('nombre_usuario', '').strip()
                password = request.form.get('password', '')
                nombre_completo = request.form.get('nombre_completo', '').strip()
                email = request.form.get('email', '').strip()
                telefono = request.form.get('telefono', '').strip()
                pais = request.form.get('pais', '').strip()
                ciudad = request.form.get('ciudad', '').strip()
                moneda = request.form.get('moneda', 'USD')

                valid, msg = validate_username(nombre_usuario)
                if not valid:
                    flash(msg, 'error')
                    return render_template('auth/registrar_usuario.html')

                valid, msg = validate_password(password)
                if not valid:
                    flash(msg, 'error')
                    return render_template('auth/registrar_usuario.html')

                if obtener_usuario_por_nombre(nombre_usuario):
                    flash('El usuario ya existe', 'error')
                else:
                    foto_perfil = None
                    file = request.files.get('foto_perfil')
                    if file and file.filename:
                        foto_perfil, error = read_profile_image(file)
                        if error:
                            flash(error, 'error')
                            return render_template('auth/registrar_usuario.html')
                    password_hash = generate_password_hash(password)
                    created = crear_usuario(
                        nombre_usuario,
                        password_hash,
                        nombre_completo,
                        email,
                        telefono,
                        pais,
                        ciudad,
                        moneda,
                        foto_perfil
                    )
                    if not created:
                        flash('Error al crear el usuario. Revisa la configuración del servidor.', 'error')
                        return render_template('auth/registrar_usuario.html')
                    flash('Usuario registrado correctamente', 'success')
                    return redirect(url_for('main.login'))
            except Exception:
                traceback.print_exc()
                flash('Error interno al procesar el registro', 'error')
                return render_template('auth/registrar_usuario.html')
        return render_template('auth/registrar_usuario.html')

    @bp.route('/perfil', methods=['GET', 'POST'])
    @login_required_custom
    def perfil():
        if 'usuario_id' not in session:
            return redirect(url_for('main.login'))
        
        usuario = get_current_usuario(refresh=True)
        if usuario is None:
            session.clear()
            return redirect(url_for('main.login'))
        
        if request.method == 'POST':
            # Actualizar Foto si se envió
            if 'foto_perfil' in request.files and request.files['foto_perfil'].filename:
                file = request.files['foto_perfil']
                if file:
                    foto_bytes, error = read_profile_image(file)
                    if error:
                        flash(error, 'error')
                        return render_template('dashboard/perfil.html', usuario=usuario)
                    actualizar_foto_perfil_db(usuario['id'], foto_bytes)
            
            # Actualizar datos personales
            nombre_completo = request.form.get('nombre_completo', '').strip()
            email = request.form.get('email', '').strip()
            telefono = request.form.get('telefono', '').strip()
            pais = request.form.get('pais', '').strip()
            ciudad = request.form.get('ciudad', '').strip()
            moneda = request.form.get('moneda', 'USD')
            
            resultado = actualizar_perfil_usuario_db(usuario['id'], nombre_completo, email, telefono, pais, ciudad, moneda)
            
            if resultado:
                flash('Perfil actualizado exitosamente', 'success')
            else:
                flash('Error al actualizar el perfil', 'error')
            
            # Recargar datos del usuario DESPUÉS de actualizar
            usuario = get_current_usuario(refresh=True)
            return redirect(url_for('main.perfil'))
        
        return render_template('dashboard/perfil.html', usuario=usuario)

    @bp.route('/editar_perfil', methods=['GET', 'POST'])
    def editar_perfil():
        if 'usuario_id' not in session:
            return redirect(url_for('main.login'))
        
        usuario = get_current_usuario(refresh=True)
        if usuario is None:
            session.clear()
            return redirect(url_for('main.login'))
        
        if request.method == 'POST':
            # Verificar si es solo actualización de foto (sin contraseña)
            if 'foto_perfil' in request.files and request.files['foto_perfil'].filename:
                file = request.files['foto_perfil']
                if file:
                    foto_bytes, error = read_profile_image(file)
                    if error:
                        flash(error, 'error')
                        return render_template('dashboard/editar_perfil.html', usuario=usuario)
                    actualizar_foto_perfil_db(usuario['id'], foto_bytes)

            password_actual = request.form.get('password_actual', '')
            password_nueva = request.form.get('password_nueva', '')
            password_confirmar = request.form.get('password_confirmar', '')
            moneda = request.form.get('moneda', 'USD')
            nombre_completo = request.form.get('nombre_completo', '')
            email = request.form.get('email', '')
            telefono = request.form.get('telefono', '')
            pais = request.form.get('pais', '')
            ciudad = request.form.get('ciudad', '')
            
            if not check_password_hash(usuario['password_hash'], password_actual):
                flash('La contraseña actual es incorrecta para aplicar cambios', 'error')
            else:
                # Actualizar datos personales y moneda
                actualizar_perfil_usuario_db(usuario['id'], nombre_completo, email, telefono, pais, ciudad, moneda)
                
                # Actualizar contraseña si se proporcionó
                if password_nueva:
                    valid, msg = validate_password(password_nueva)
                    if not valid:
                        flash(msg, 'error')
                    elif password_nueva != password_confirmar:
                        flash('Las nuevas contraseñas no coinciden', 'error')
                    else:
                        nueva_password_hash = generate_password_hash(password_nueva)
                        actualizar_password_usuario(usuario['id'], nueva_password_hash)
                        flash('Perfil y contraseña actualizados exitosamente', 'success')
                else:
                    flash('Preferencias actualizadas correctamente', 'success')
                
                # Recargar datos del usuario
                usuario = get_current_usuario(refresh=True)
                return redirect(url_for('main.editar_perfil'))
        
        return render_template('dashboard/editar_perfil.html', usuario=usuario)

    @bp.route('/editar_registro', methods=['GET', 'POST'])
    @login_required_custom
    def editar_registro():
        usuario_id = session['usuario_id']
        usuario = get_current_usuario()
        
        if request.method == 'GET':
            registro_id = request.args.get('id')
            tipo = request.args.get('tipo', '').strip()
            
            if not registro_id or not tipo:
                return redirect(url_for('main.registrar'))
            
            registro = None
            if tipo == 'ingreso':
                registro = obtener_ingreso_por_id(registro_id, usuario_id)
            elif tipo == 'gasto':
                registro = obtener_gasto_por_id(registro_id, usuario_id)
            elif tipo == 'inversion':
                registro = obtener_inversion_por_id(registro_id, usuario_id)
            
            if not registro:
                flash('Registro no encontrado', 'error')
                return redirect(url_for('main.registrar'))
            
            categorias = obtener_categorias(usuario_id)
            return render_template(
                'forms/editar_registro.html',
                usuario=usuario,
                registro=registro,
                tipo=tipo,
                categorias=categorias,
                return_to=request.args.get('return_to', ''),
                return_query=request.args.get('return_query', ''),
            )
        
        elif request.method == 'POST':
            registro_id = request.form.get('registro_id')
            tipo = request.form.get('tipo', '').strip()
            monto = request.form.get('monto', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            fecha = request.form.get('fecha', '').strip()
            
            valid, msg = validate_tipo(tipo)
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.registrar'))
            
            valid, msg = validate_monto(monto)
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.registrar'))
            
            valid, msg = validate_fecha(fecha)
            if not valid:
                flash(msg, 'error')
                return redirect(url_for('main.registrar'))
            
            categoria_id = request.form.get('categoria_id') or None
            if categoria_id:
                try:
                    categoria_id = int(categoria_id)
                except ValueError:
                    categoria_id = None

            # Validar que la categoría (si existe) sea del tipo correcto
            if categoria_id:
                valid_cat, msg_cat = validar_categoria_para_tipo(usuario_id, categoria_id, tipo)
                if not valid_cat:
                    flash(msg_cat, 'error')
                    return redirect(url_for('main.editar_registro', id=registro_id, tipo=tipo))

            if tipo == 'ingreso':
                actualizar_ingreso(registro_id, usuario_id, monto, descripcion, fecha, categoria_id)
            elif tipo == 'gasto':
                actualizar_gasto(registro_id, usuario_id, monto, descripcion, fecha, categoria_id)
            elif tipo == 'inversion':
                actualizar_inversion(registro_id, usuario_id, monto, descripcion, fecha, categoria_id)
            
            destino = request.form.get('return_to', '')
            if destino == 'buscar':
                return _redirect_buscar(request.form.get('return_query', ''))
            
            # Redirigir según el tipo de registro
            if tipo == 'inversion':
                return redirect(url_for('main.inversiones'))
            elif tipo == 'gasto':
                return redirect(url_for('main.registrar'))
            else:  # ingreso
                return redirect(url_for('main.registrar'))

    @bp.route('/eliminar_registro', methods=['POST'])
    @login_required_custom
    def eliminar_registro():
        registro_id = request.form.get('registro_id')
        tipo = request.form.get('tipo', '').strip()
        usuario_id = session['usuario_id']
        
        valid, msg = validate_tipo(tipo)
        if not valid:
            flash(msg, 'error')
            return redirect(url_for('main.registrar'))
        
        if tipo == 'ingreso':
            eliminar_ingreso(registro_id, usuario_id)
        elif tipo == 'gasto':
            eliminar_gasto(registro_id, usuario_id)
        elif tipo == 'inversion':
            eliminar_inversion(registro_id, usuario_id)
        
        flash('Registro eliminado exitosamente', 'success')
        destino = request.form.get('return_to', 'registrar')
        if destino == 'buscar':
            return _redirect_buscar(request.form.get('return_query', ''))
        return redirect(url_for('main.registrar'))

    @bp.route('/exportar_pdf', methods=['GET'])
    @login_required_custom
    def exportar_pdf():
        mes = request.args.get('mes', datetime.now().strftime('%Y-%m'))
        usuario = get_current_usuario()
        if usuario is None:
            session.clear()
            return redirect(url_for('main.login'))
        currency_code = usuario.get('moneda', 'USD')
        
        # Obtener datos del mes
        anio, mes_num = (int(mes[:4]), int(mes[5:]))
        
        ingresos_mes = obtener_ingresos_por_mes(usuario['id'], anio, mes_num)
        gastos_mes = obtener_gastos_por_mes(usuario['id'], anio, mes_num)
        inversiones_mes = obtener_inversiones_por_mes(usuario['id'], anio, mes_num)
        
        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch, 
                               title="Reporte financiero", author=usuario['nombre_usuario'])
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=22, textColor=colors.HexColor('#1a1a1a'), spaceAfter=24, alignment=TA_CENTER, fontName='Helvetica-Bold')
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#555555'), spaceAfter=6, alignment=TA_LEFT)
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#1a1a1a'), spaceAfter=10, spaceBefore=12, fontName='Helvetica-Bold', alignment=TA_LEFT)

        def formatear_fecha_reporte(fecha):
            if isinstance(fecha, datetime):
                return fecha.strftime('%d/%m/%Y')
            if isinstance(fecha, date):
                return fecha.strftime('%d/%m/%Y')
            if isinstance(fecha, str):
                try:
                    return datetime.fromisoformat(fecha).strftime('%d/%m/%Y')
                except ValueError:
                    try:
                        return datetime.strptime(fecha, '%Y-%m-%d').strftime('%d/%m/%Y')
                    except ValueError:
                        return fecha
            return str(fecha)

        def formatear_moneda_reporte(valor):
            try:
                valor_float = float(valor)
            except (TypeError, ValueError):
                return str(valor)

            valor_formateado = f"{valor_float:,.2f}"
            parte_entera, parte_decimal = valor_formateado.rsplit('.', 1)
            parte_entera = parte_entera.replace(',', '.')
            valor_formateado = f"{parte_entera},{parte_decimal}"

            if currency_code == 'COP':
                return f"$ {valor_formateado}"
            return f"{currency_code} {valor_formateado}"
        
        # Obtener fecha actual formateada
        hoy = datetime.now()
        dias = hoy.strftime('%d')
        meses_es = {'01': 'enero', '02': 'febrero', '03': 'marzo', '04': 'abril', '05': 'mayo', '06': 'junio', '07': 'julio', '08': 'agosto', '09': 'septiembre', '10': 'octubre', '11': 'noviembre', '12': 'diciembre'}
        mes_es = meses_es[hoy.strftime('%m')]
        año = hoy.strftime('%Y')
        fecha_str = f"{dias} de {mes_es} de {año}"
        
        # Título y datos generales
        title = Paragraph("Reporte financiero", title_style)
        elements.append(title)
        
        subtitle1 = Paragraph(f"Fecha: {fecha_str}", subtitle_style)
        elements.append(subtitle1)
        subtitle2 = Paragraph(f"Usuario: {usuario['nombre_usuario']}", subtitle_style)
        elements.append(subtitle2)
        elements.append(Spacer(1, 0.25*inch))
        
        # Función para crear tabla
        def crear_tabla(registros, titulo):
            if not registros:
                elements.append(Paragraph(titulo, heading_style))
                elements.append(Paragraph("<i>No hay registros</i>", styles['Normal']))
                elements.append(Spacer(1, 0.2*inch))
                return
            
            elements.append(Paragraph(titulo, heading_style))
            
            data = [['Fecha', 'Descripción', 'Monto']]
            total = 0
            for reg in registros:
                fecha = formatear_fecha_reporte(reg['fecha'])
                data.append([fecha, reg['descripcion'], formatear_moneda_reporte(reg['monto'])])
                total += reg['monto']
            
            # Agregar fila de total sin HTML
            data.append(['', 'Total', formatear_moneda_reporte(total)])
            
            table = Table(data, colWidths=[1.2*inch, 3*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8e8e8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8e8e8')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('TOPPADDING', (0, -1), (-1, -1), 12),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Agregar tablas
        crear_tabla(ingresos_mes, "Ingresos")
        crear_tabla(gastos_mes, "Gastos")
        crear_tabla(inversiones_mes, "Inversiones")
        
        
        # Generar PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f"Reporte_{mes}.pdf")

    @bp.route('/exportar_excel')
    @login_required_custom
    def exportar_excel():
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        mes = request.args.get('mes', datetime.now().strftime('%Y-%m'))
        usuario = get_current_usuario()
        if usuario is None:
            session.clear()
            return redirect(url_for('main.login'))
        currency_code = usuario.get('moneda', 'USD')
        
        # Obtener datos
        anio, mes_num = (int(mes[:4]), int(mes[5:]))
        ingresos = obtener_ingresos_por_mes(usuario['id'], anio, mes_num)
        gastos = obtener_gastos_por_mes(usuario['id'], anio, mes_num)
        inversiones = obtener_inversiones_por_mes(usuario['id'], anio, mes_num)
        
        # Crear workbook
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte'
        
        # Estilos
        title_font = Font(name='Calibri', size=22, bold=True)
        subtitle_font = Font(name='Calibri', size=11, color='555555')
        header_font = Font(name='Calibri', size=13, bold=True)
        data_font = Font(name='Calibri', size=11)
        total_font = Font(name='Calibri', size=11, bold=True)
        
        header_fill = PatternFill(start_color='e8e8e8', end_color='e8e8e8', fill_type='solid')
        total_fill = PatternFill(start_color='e8e8e8', end_color='e8e8e8', fill_type='solid')
        
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        alignment_right = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='cccccc'),
            right=Side(style='thin', color='cccccc'),
            top=Side(style='thin', color='cccccc'),
            bottom=Side(style='thin', color='cccccc')
        )

        def formatear_fecha_reporte(fecha):
            if isinstance(fecha, datetime):
                return fecha.strftime('%d/%m/%Y')
            if isinstance(fecha, date):
                return fecha.strftime('%d/%m/%Y')
            if isinstance(fecha, str):
                try:
                    return datetime.fromisoformat(fecha).strftime('%d/%m/%Y')
                except ValueError:
                    try:
                        return datetime.strptime(fecha, '%Y-%m-%d').strftime('%d/%m/%Y')
                    except ValueError:
                        return fecha
            return str(fecha)
        
        # Fecha formateada
        hoy = datetime.now()
        meses_es = {'01': 'enero', '02': 'febrero', '03': 'marzo', '04': 'abril', '05': 'mayo', '06': 'junio', '07': 'julio', '08': 'agosto', '09': 'septiembre', '10': 'octubre', '11': 'noviembre', '12': 'diciembre'}
        mes_es = meses_es[hoy.strftime('%m')]
        fecha_str = f"{hoy.strftime('%d')} de {mes_es} de {hoy.strftime('%Y')}"
        
        # Agregar títulos
        row = 1
        ws[f'A{row}'] = 'Reporte financiero'
        ws[f'A{row}'].font = title_font
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].alignment = alignment_center
        ws.row_dimensions[row].height = 35
        
        row += 2
        ws[f'A{row}'] = f'Fecha: {fecha_str}'
        ws[f'A{row}'].font = subtitle_font
        
        row += 1
        ws[f'A{row}'] = f'Usuario: {usuario["nombre_usuario"]}'
        ws[f'A{row}'].font = subtitle_font

        row += 1
        ws[f'A{row}'] = f'Moneda: {currency_code}'
        ws[f'A{row}'].font = subtitle_font
        
        row += 2

        def formato_moneda_excel(valor):
            try:
                float(valor)
            except (TypeError, ValueError):
                return str(valor)

            if currency_code == 'COP':
                return '"$" #,##0.00'
            return f'"{currency_code}" #,##0.00'
        
        # Función para agregar tabla
        def agregar_tabla(registros, titulo, inicio_row):
            nonlocal row
            row = inicio_row
            
            ws[f'A{row}'] = titulo
            ws[f'A{row}'].font = header_font
            row += 1
            
            if not registros:
                ws[f'A{row}'] = 'No hay registros'
                row += 1
                return row
            
            # Encabezados
            ws[f'A{row}'] = 'Fecha'
            ws[f'B{row}'] = 'Descripción'
            ws[f'C{row}'] = 'Monto'
            
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = thin_border
            
            row += 1
            
            # Datos
            total = 0
            for reg in registros:
                fecha = formatear_fecha_reporte(reg['fecha'])
                
                ws[f'A{row}'] = fecha
                ws[f'B{row}'] = reg['descripcion']
                ws[f'C{row}'] = reg['monto']
                
                ws[f'A{row}'].font = data_font
                ws[f'B{row}'].font = data_font
                ws[f'C{row}'].font = data_font
                
                ws[f'A{row}'].alignment = alignment_center
                ws[f'B{row}'].alignment = alignment_left
                ws[f'C{row}'].alignment = alignment_right

                ws[f'C{row}'].number_format = formato_moneda_excel(reg['monto'])
                
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = thin_border
                
                total += reg['monto']
                row += 1
            
            # Fila de total
            ws[f'B{row}'] = 'Total'
            ws[f'C{row}'] = total
            
            ws[f'B{row}'].font = total_font
            ws[f'C{row}'].font = total_font
            
            ws[f'B{row}'].alignment = alignment_right
            ws[f'C{row}'].alignment = alignment_right
            
            ws[f'C{row}'].number_format = formato_moneda_excel(total)
            
            ws[f'B{row}'].fill = total_fill
            ws[f'C{row}'].fill = total_fill
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = thin_border
            
            row += 2
            return row
        
        # Agregar tablas
        row = agregar_tabla(ingresos, 'Ingresos', row)
        row = agregar_tabla(gastos, 'Gastos', row)
        row = agregar_tabla(inversiones, 'Inversiones', row)
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        
        # Guardar en buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                         as_attachment=True, download_name=f"Reporte_{mes}.xlsx")


    @bp.route('/metas', methods=['GET', 'POST'])
    @login_required_custom
    def metas():
        usuario_id = session['usuario_id']
        if request.method == 'POST':
            nombre = request.form.get('nombre')
            monto_objetivo = request.form.get('monto_objetivo')
            if agregar_meta(usuario_id, nombre, monto_objetivo):
                flash('Meta creada', 'success')
            else:
                flash('Error al crear meta', 'error')
            return redirect(url_for('main.metas'))
            
        mis_metas = obtener_metas(usuario_id)
        for m in mis_metas:
            # Convertir ambos a float para evitar errores de tipos
            m['monto_actual'] = float(m.get('monto_actual') or 0)
            m['monto_objetivo'] = float(m.get('monto_objetivo') or 0)
            m['porcentaje'] = round((m['monto_actual'] / m['monto_objetivo'] * 100), 1) if m['monto_objetivo'] > 0 else 0
            
        return render_template('config/metas.html', metas=mis_metas)

    @bp.route('/eliminar_meta/<int:id>')
    @login_required_custom
    def eliminar_meta(id):
        if eliminar_meta_db(id, session['usuario_id']):
            flash('Meta eliminada', 'success')
        return redirect(url_for('main.metas'))

    @bp.route('/aporte_meta/<int:id>', methods=['POST'])
    @login_required_custom
    def aporte_meta(id):
        import json
        data = json.loads(request.data)
        monto = float(data.get('monto', 0))
        
        if monto <= 0:
            return jsonify({'success': False, 'message': 'Monto inválido'})
        
        if actualizar_aporte_meta(id, session['usuario_id'], monto):
            return jsonify({'success': True, 'message': 'Aporte registrado'})
        else:
            return jsonify({'success': False, 'message': 'Error al registrar aporte'})


    @bp.route('/recurrentes', methods=['GET', 'POST'])
    @login_required_custom
    def recurrentes():
        usuario_id = session['usuario_id']
        if request.method == 'POST':
            tipo = request.form.get('tipo')
            monto = request.form.get('monto')
            descripcion = request.form.get('descripcion')
            frecuencia = request.form.get('frecuencia')
            proxima_fecha = request.form.get('proxima_fecha')
            edit_id = request.form.get('edit_id')

            if edit_id:
                try:
                    edit_id_int = int(edit_id)
                except ValueError:
                    edit_id_int = None
                if edit_id_int and actualizar_recurrente(edit_id_int, usuario_id, tipo, monto, descripcion, frecuencia, proxima_fecha):
                    flash('Programación actualizada', 'success')
                else:
                    flash('Error al actualizar la programación', 'error')
            else:
                if agregar_recurrente(usuario_id, tipo, monto, descripcion, frecuencia, proxima_fecha):
                    flash('Transacción recurrente programada', 'success')
                else:
                    flash('Error al programar transacción', 'error')

            return redirect(url_for('main.recurrentes'))
            
        mis_recurrentes = obtener_recurrentes(usuario_id)
        return render_template('config/recurrentes.html', recurrentes=mis_recurrentes)

    @bp.route('/eliminar_recurrente/<int:id>')
    @login_required_custom
    def eliminar_recurrente(id):
        if eliminar_recurrente_db(id, session['usuario_id']):
            flash('Programación eliminada', 'success')
        return redirect(url_for('main.recurrentes'))

    @bp.route('/seguridad', methods=['GET', 'POST'])
    @login_required_custom
    def seguridad():
        try:
            usuario_id = session['usuario_id']
            usuario = get_current_usuario(refresh=True)
            
            if request.method == 'POST':
                # Checkbox state
                MFA = request.form.get('MFA') is not None
                
                # Si se está activando (y no estaba activo)
                if MFA and not usuario.get('MFA'):
                    try:
                        import pyotp
                        import qrcode
                        import io
                        import base64
                    except ImportError as e:
                        flash("Error: Faltan librerías de seguridad (pyotp/qrcode)", "error")
                        return redirect(url_for('main.seguridad'))

                    # Generar nuevo secreto
                    secret = pyotp.random_base32()
                    if not actualizar_totp_secret_db(usuario_id, secret):
                        flash("Error al guardar secreto de seguridad en base de datos", "error")
                        return redirect(url_for('main.seguridad'))
                    
                    # Generar QR
                    uri = pyotp.totp.TOTP(secret).provisioning_uri(name=usuario['nombre_usuario'], issuer_name="App Money")
                    img = qrcode.make(uri)
                    buffered = io.BytesIO()
                    img.save(buffered, format="PNG")
                    qr_b64 = base64.b64encode(buffered.getvalue()).decode()
                    
                    # Activar en BD
                    actualizar_seguridad_usuario(usuario_id, True)
                    
                    # Recargar usuario para verificar
                    usuario_actualizado = get_current_usuario(refresh=True)
                    
                    # Renderizar con el código QR
                    return render_template('auth/seguridad.html', usuario=usuario_actualizado, qr_code=qr_b64, secret=secret)
                
                elif not MFA and usuario.get('MFA'):
                    # Desactivar
                    actualizar_seguridad_usuario(usuario_id, False)
                    return redirect(url_for('main.seguridad'))
                
                # Si no hubo cambio de estado pero se hizo post
                return redirect(url_for('main.seguridad'))
                
            return render_template('auth/seguridad.html', usuario=usuario)
        except Exception as e:
            import traceback
            traceback.print_exc()
            flash(f"Error interno: {str(e)}", "error")
            return redirect(url_for('main.dashboard')) # Fallback seguro

    @bp.route('/buscar')
    @login_required_custom
    def buscar():
        usuario_id = session['usuario_id']
        filtros = _filtros_desde_request(request.args)
        try:
            pagina = max(1, int(request.args.get('page', 1)))
        except (TypeError, ValueError):
            pagina = 1

        datos = explorar_transacciones(usuario_id, filtros, pagina=pagina, por_pagina=25)
        categorias = obtener_categorias(usuario_id)

        return render_template(
            'dashboard/buscar.html',
            resultados=datos['resultados'],
            resumen=datos['resumen'],
            paginacion=datos['paginacion'],
            filtros=filtros,
            categorias=categorias,
        )

    @bp.route('/buscar/exportar')
    @login_required_custom
    def exportar_buscar():
        usuario = get_current_usuario()
        if usuario is None:
            session.clear()
            return redirect(url_for('main.login'))
        currency_code = usuario.get('moneda', 'USD')
        usuario_id = session['usuario_id']
        filtros = _filtros_desde_request(request.args)
        datos = obtener_transacciones_exportacion(usuario_id, filtros)
        filas = []
        for t in datos:
            filas.append({
                'Fecha': _formatear_fecha_ddmmaa(t.get('fecha', '')),
                'Descripción': t.get('descripcion', '') or '',
                'Tipo': t.get('tipo', ''),
                'Categoría': t.get('categoria_nombre') or 'Sin categoría',
                'Monto': t.get('monto', ''),
            })

        df = pd.DataFrame(filas, columns=['Fecha', 'Descripción', 'Tipo', 'Categoría', 'Monto'])
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Transacciones', startrow=3)
            worksheet = writer.book['Transacciones']
            worksheet.sheet_view.showGridLines = False

            # Título y subtítulo del informe
            total_resultados = len(df)
            titulo = 'Informe de transacciones'
            
            # Mapear filtros a texto legible
            tipo_texto = 'Todos'
            if filtros.get('tipo') == 'ingreso':
                tipo_texto = 'Ingreso'
            elif filtros.get('tipo') == 'gasto':
                tipo_texto = 'Gasto'
            elif filtros.get('tipo') == 'inversion':
                tipo_texto = 'Inversión'
            
            categoria_texto = 'Todas'
            if filtros.get('categoria_id'):
                # Obtener nombre de la categoría
                conn = get_connection()
                if conn:
                    try:
                        cursor = conn.cursor(dictionary=True)
                        cursor.execute('SELECT nombre FROM categorias WHERE id = %s', (filtros.get('categoria_id'),))
                        cat = cursor.fetchone()
                        if cat:
                            categoria_texto = cat['nombre']
                        cursor.close()
                    finally:
                        conn.close()
            
            subtitulo = (
                f"Rango: {_formatear_fecha_ddmmaa(filtros.get('desde')) if filtros.get('desde') else 'todos'} → {_formatear_fecha_ddmmaa(filtros.get('hasta')) if filtros.get('hasta') else 'todos'} | "
                f"Tipo: {tipo_texto} | "
                f"Categoría: {categoria_texto} | "
                f"Movimientos: {total_resultados}"
            )
            
            # Fila 1: Título
            worksheet['A1'] = titulo
            worksheet.merge_cells('A1:E1')
            worksheet['A1'].font = Font(size=14, bold=True, color='000000')
            worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet.row_dimensions[1].height = 25
            
            # Fila 2: Subtítulo
            worksheet['A2'] = subtitulo
            worksheet.merge_cells('A2:E2')
            worksheet['A2'].font = Font(size=10, italic=True, color='000000')
            worksheet['A2'].alignment = Alignment(horizontal='left', vertical='center')
            worksheet.row_dimensions[2].height = 20

            worksheet['A3'] = f"Moneda: {currency_code}"
            worksheet['A3'].font = Font(size=10, italic=True, color='000000')
            worksheet['A3'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Fila 3: Encabezados
            thin_black = Side(style='thin', color='000000')
            header_border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
            
            for col_num, header in enumerate(['Fecha', 'Descripción', 'Tipo', 'Categoría', 'Monto'], 1):
                cell = worksheet.cell(row=4, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color='000000', size=11)
                cell.fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
                cell.border = header_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            worksheet.row_dimensions[4].height = 20

            # Formatear las filas de datos
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=1, max_col=5), start=5):
                for cell in row:
                    cell.border = header_border
                    cell.alignment = Alignment(vertical='center')

                # Fecha (columna A): DD/MM/YYYY
                try:
                    raw_fecha = row[0].value
                    if raw_fecha:
                        if isinstance(raw_fecha, str):
                            try:
                                parsed = datetime.fromisoformat(raw_fecha)
                            except Exception:
                                parsed = datetime.strptime(raw_fecha, '%Y-%m-%d')
                            row[0].value = parsed.date()
                        row[0].number_format = 'DD/MM/YYYY'
                        row[0].alignment = Alignment(horizontal='center', vertical='center')
                except Exception:
                    pass

                # Descripción (columna B): sin cambios
                try:
                    row[1].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                except Exception:
                    pass

                # Tipo (columna C): capitalizar
                try:
                    if row[2].value:
                        row[2].value = str(row[2].value).capitalize()
                        row[2].alignment = Alignment(horizontal='left', vertical='center')
                except Exception:
                    pass

                # Categoría (columna D): capitalizar
                try:
                    if row[3].value:
                        row[3].value = str(row[3].value)
                        row[3].alignment = Alignment(horizontal='left', vertical='center')
                except Exception:
                    pass

                # Monto (columna E): formato numérico con separadores de miles
                try:
                    if row[4].value is not None and row[4].value != '':
                        monto_val = row[4].value
                        if isinstance(monto_val, str):
                            monto_val = float(str(monto_val).replace('.', '').replace(',', '.'))
                        row[4].value = float(monto_val)
                        row[4].number_format = '"$" #,##0.00' if currency_code == 'COP' else f'"{currency_code}" #,##0.00'
                        row[4].alignment = Alignment(horizontal='right', vertical='center')
                except Exception:
                    pass

            # Anchos de columnas
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 22
            worksheet.column_dimensions['E'].width = 16
            
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'transacciones_{datetime.now().strftime("%Y%m%d")}.xlsx',
        )


    @bp.route('/categorias', methods=['GET', 'POST'])
    @login_required_custom
    def categorias_route():
        usuario_id = session['usuario_id']
        if request.method == 'POST':
            nombre = request.form.get('nombre')
            tipo = request.form.get('tipo')
            color = request.form.get('color')
            if agregar_categoria(usuario_id, nombre, tipo, color):
                flash('Categoría agregada', 'success')
            else:
                flash('Error al agregar categoría', 'error')
            return redirect(url_for('main.categorias_route'))
        
        categorias_list = obtener_categorias(usuario_id)
        return render_template('config/categorias.html', categorias=categorias_list)

    @bp.route('/eliminar_categoria/<int:id>')
    @login_required_custom
    def eliminar_categoria_route(id):
        if eliminar_categoria_db(id, session['usuario_id']):
            flash('Categoría eliminada', 'success')
        return redirect(url_for('main.categorias_route'))

    @bp.route('/presupuestos', methods=['GET', 'POST'])
    @login_required_custom
    def presupuestos():
        usuario_id = session['usuario_id']
        if request.method == 'POST':
            categoria_id = request.form.get('categoria_id')
            try:
                monto = float(request.form.get('monto', 0))
            except (ValueError, TypeError):
                monto = 0
            if agregar_o_actualizar_presupuesto(usuario_id, categoria_id, monto):
                flash('Presupuesto actualizado', 'success')
            else:
                flash('Error al actualizar presupuesto', 'error')
            return redirect(url_for('main.presupuestos'))
            
        mis_presupuestos = obtener_presupuestos_simples(usuario_id)
        return render_template('config/presupuestos.html', presupuestos=mis_presupuestos)

    @bp.route('/editar_presupuesto/<int:id>', methods=['POST'])
    @login_required_custom
    def editar_presupuesto(id):
        usuario_id = session['usuario_id']
        try:
            gastado = float(request.form.get('gastado', 0))
        except (ValueError, TypeError):
            gastado = 0
        
        if actualizar_presupuesto_gastado(usuario_id, id, gastado):
            # Si es una petición AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True, 'gastado': gastado})
            flash('Presupuesto actualizado', 'success')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'Error al actualizar presupuesto'})
            flash('Error al actualizar presupuesto', 'error')
        
        return redirect(url_for('main.presupuestos'))

    @bp.route('/eliminar_presupuesto/<int:id>', methods=['GET'])
    @login_required_custom
    def eliminar_presupuesto(id):
        usuario_id = session['usuario_id']
        
        if eliminar_presupuesto_db(usuario_id, id):
            flash('Presupuesto eliminado', 'success')
        else:
            flash('Error al eliminar presupuesto', 'error')
        
        return redirect(url_for('main.presupuestos'))

    app.register_blueprint(bp)

# --- Funciones auxiliares para la lógica de base de datos ---
def obtener_usuario_por_id(usuario_id):
    conn = get_connection()
    if conn is None:
        return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            'SELECT id, nombre_usuario, password_hash, totp_secret, MFA, ultimo_login, nombre_completo, '
            'email, telefono, pais, ciudad, moneda, '
            'CASE WHEN foto_perfil IS NOT NULL THEN 1 ELSE 0 END AS foto_perfil, creado_en '
            'FROM usuarios WHERE id = %s',
            (usuario_id,)
        )
        usuario = cursor.fetchone()
        cursor.close()
        return usuario
    finally:
        conn.close()

def obtener_usuario_por_nombre(nombre_usuario):
    conn = get_connection()
    if conn is None:
        return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            'SELECT id, nombre_usuario, password_hash, totp_secret, MFA, ultimo_login, nombre_completo, '
            'email, telefono, pais, ciudad, moneda, '
            'CASE WHEN foto_perfil IS NOT NULL THEN 1 ELSE 0 END AS foto_perfil, creado_en '
            'FROM usuarios WHERE nombre_usuario = %s',
            (nombre_usuario,)
        )
        usuario = cursor.fetchone()
        cursor.close()
        return usuario
    finally:
        conn.close()

def crear_usuario(
    nombre_usuario,
    password_hash,
    nombre_completo='',
    email='',
    telefono='',
    pais='',
    ciudad='',
    moneda='USD',
    foto_perfil=None
):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO usuarios (nombre_usuario, password_hash, nombre_completo, email, telefono, pais, ciudad, moneda, foto_perfil) '
            'VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)',
            (nombre_usuario, password_hash, nombre_completo, email, telefono, pais, ciudad, moneda, foto_perfil)
        )
        usuario_id = cursor.lastrowid
        
        # Crear categorías por defecto
        categorias_defecto = [
            ('Gastos Generales', 'gasto', '#dc2626'),
            ('Ingresos Generales', 'ingreso', '#16a34a'),
            ('Inversiones Generales', 'inversion', '#2563eb')
        ]
        
        for nombre, tipo, color in categorias_defecto:
            cursor.execute('INSERT INTO categorias (usuario_id, nombre, tipo, color) VALUES (%s, %s, %s, %s)',
                          (usuario_id, nombre, tipo, color))
        
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def obtener_ingresos(usuario_id):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, monto, descripcion, fecha, categoria_id, creado_en FROM ingresos WHERE usuario_id = %s ORDER BY fecha DESC, id DESC', (usuario_id,))
        ingresos = cursor.fetchall()
        cursor.close()
        return ingresos
    finally:
        conn.close()

def obtener_gastos(usuario_id):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, monto, descripcion, fecha, categoria_id, creado_en FROM gastos WHERE usuario_id = %s ORDER BY fecha DESC, id DESC', (usuario_id,))
        gastos = cursor.fetchall()
        cursor.close()
        return gastos
    finally:
        conn.close()

def obtener_categoria_defecto(usuario_id, tipo):
    """Obtiene la categoría por defecto para un tipo de transacción"""
    conn = get_connection()
    if conn is None:
        return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id FROM categorias WHERE usuario_id = %s AND tipo = %s LIMIT 1', (usuario_id, tipo))
        resultado = cursor.fetchone()
        cursor.close()
        return resultado['id'] if resultado else None
    finally:
        conn.close()

def agregar_ingreso(usuario_id, monto, descripcion, fecha, categoria_id=None):
    conn = get_connection()
    if conn is None:
        return False
    try:
        if not categoria_id:
            categoria_id = obtener_categoria_defecto(usuario_id, 'ingreso')
        cursor = conn.cursor()
        cursor.execute('INSERT INTO ingresos (usuario_id, monto, descripcion, fecha, categoria_id) VALUES (%s, %s, %s, %s, %s)', (usuario_id, monto, descripcion, fecha, categoria_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()
def agregar_inversion(usuario_id, tipo, monto, descripcion, fecha, categoria_id=None):
    conn = get_connection()
    if conn is None:
        return False
    try:
        if not categoria_id:
            categoria_id = obtener_categoria_defecto(usuario_id, 'inversion')
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO inversiones (usuario_id, tipo, monto, descripcion, fecha, categoria_id)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (usuario_id, tipo, monto, descripcion, fecha, categoria_id),
        )
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        print(f"agregar_inversion: {e}")
        return False
    finally:
        conn.close()

def obtener_inversiones(usuario_id):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, tipo, monto, descripcion, fecha, categoria_id, creado_en FROM inversiones WHERE usuario_id = %s ORDER BY fecha DESC, id DESC', (usuario_id,))
        inversiones = cursor.fetchall()
        cursor.close()
        return inversiones
    finally:
        conn.close()

def obtener_inversiones_por_mes(usuario_id, anio, mes):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        desde, hasta = calcular_rango_mes(anio, mes)
        cursor.execute('SELECT id, usuario_id, tipo, monto, descripcion, fecha, categoria_id, creado_en FROM inversiones WHERE usuario_id = %s AND fecha BETWEEN %s AND %s ORDER BY fecha DESC, id DESC', (usuario_id, desde, hasta))
        inversiones = cursor.fetchall()
        cursor.close()
        return inversiones
    finally:
        conn.close()
def agregar_gasto(usuario_id, monto, descripcion, fecha, categoria_id=None):
    """Agrega un gasto - categoria_id es opcional"""
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO gastos (usuario_id, monto, descripcion, fecha, categoria_id) VALUES (%s, %s, %s, %s, %s)', 
                      (usuario_id, monto, descripcion, fecha, categoria_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def obtener_ingresos_por_mes(usuario_id, anio, mes):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        desde, hasta = calcular_rango_mes(anio, mes)
        cursor.execute('SELECT id, usuario_id, monto, descripcion, fecha, categoria_id, creado_en FROM ingresos WHERE usuario_id = %s AND fecha BETWEEN %s AND %s ORDER BY fecha DESC, id DESC', (usuario_id, desde, hasta))
        ingresos = cursor.fetchall()
        cursor.close()
        return ingresos
    finally:
        conn.close()

def obtener_gastos_por_mes(usuario_id, anio, mes):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        desde, hasta = calcular_rango_mes(anio, mes)
        cursor.execute('SELECT id, usuario_id, monto, descripcion, fecha, categoria_id, creado_en FROM gastos WHERE usuario_id = %s AND fecha BETWEEN %s AND %s ORDER BY fecha DESC, id DESC', (usuario_id, desde, hasta))
        gastos = cursor.fetchall()
        cursor.close()
        return gastos
    finally:
        conn.close()
def actualizar_password_usuario(usuario_id, nueva_password_hash):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE usuarios SET password_hash = %s WHERE id = %s', (nueva_password_hash, usuario_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def eliminar_ingreso(ingreso_id, usuario_id):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM ingresos WHERE id = %s AND usuario_id = %s', (ingreso_id, usuario_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def eliminar_gasto(gasto_id, usuario_id):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM gastos WHERE id = %s AND usuario_id = %s', (gasto_id, usuario_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def eliminar_inversion(inversion_id, usuario_id):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM inversiones WHERE id = %s AND usuario_id = %s', (inversion_id, usuario_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def obtener_ingreso_por_id(ingreso_id, usuario_id):
    conn = get_connection()
    if conn is None:
        return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, monto, descripcion, fecha, categoria_id, creado_en FROM ingresos WHERE id = %s AND usuario_id = %s', (ingreso_id, usuario_id))
        ingreso = cursor.fetchone()
        cursor.close()
        return ingreso
    finally:
        conn.close()

def obtener_gasto_por_id(gasto_id, usuario_id):
    conn = get_connection()
    if conn is None:
        return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, monto, descripcion, fecha, categoria_id, creado_en FROM gastos WHERE id = %s AND usuario_id = %s', (gasto_id, usuario_id))
        gasto = cursor.fetchone()
        cursor.close()
        return gasto
    finally:
        conn.close()

def obtener_inversion_por_id(inversion_id, usuario_id):
    conn = get_connection()
    if conn is None:
        return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, tipo, monto, descripcion, fecha, categoria_id, creado_en FROM inversiones WHERE id = %s AND usuario_id = %s', (inversion_id, usuario_id))
        inversion = cursor.fetchone()
        cursor.close()
        return inversion
    finally:
        conn.close()

def actualizar_ingreso(ingreso_id, usuario_id, monto, descripcion, fecha, categoria_id=None):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        if categoria_id is not None:
            cursor.execute('UPDATE ingresos SET monto = %s, descripcion = %s, fecha = %s, categoria_id = %s WHERE id = %s AND usuario_id = %s', 
                          (monto, descripcion, fecha, categoria_id, ingreso_id, usuario_id))
        else:
            cursor.execute('UPDATE ingresos SET monto = %s, descripcion = %s, fecha = %s WHERE id = %s AND usuario_id = %s', 
                          (monto, descripcion, fecha, ingreso_id, usuario_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def actualizar_gasto(gasto_id, usuario_id, monto, descripcion, fecha, categoria_id=None):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        if categoria_id is not None:
            cursor.execute('UPDATE gastos SET monto = %s, descripcion = %s, fecha = %s, categoria_id = %s WHERE id = %s AND usuario_id = %s', 
                          (monto, descripcion, fecha, categoria_id, gasto_id, usuario_id))
        else:
            cursor.execute('UPDATE gastos SET monto = %s, descripcion = %s, fecha = %s WHERE id = %s AND usuario_id = %s', 
                          (monto, descripcion, fecha, gasto_id, usuario_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def actualizar_inversion(inversion_id, usuario_id, monto, descripcion, fecha, categoria_id=None):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        if categoria_id is not None:
            cursor.execute('UPDATE inversiones SET monto = %s, descripcion = %s, fecha = %s, categoria_id = %s WHERE id = %s AND usuario_id = %s', 
                          (monto, descripcion, fecha, categoria_id, inversion_id, usuario_id))
        else:
            cursor.execute('UPDATE inversiones SET monto = %s, descripcion = %s, fecha = %s WHERE id = %s AND usuario_id = %s', 
                          (monto, descripcion, fecha, inversion_id, usuario_id))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def crear_token_recuperacion(usuario_id, token, expires_at):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO password_reset_tokens (usuario_id, token, expires_at) VALUES (%s, %s, %s)', 
                      (usuario_id, token, expires_at))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def obtener_token_recuperacion(token):
    conn = get_connection()
    if conn is None:
        return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, token, expires_at, captcha_attempt, verified, creado_en FROM password_reset_tokens WHERE token = %s', (token,))
        token_data = cursor.fetchone()
        cursor.close()
        return token_data
    finally:
        conn.close()

def eliminar_token_recuperacion(token):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM password_reset_tokens WHERE token = %s', (token,))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def actualizar_token_verificado(token):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE password_reset_tokens SET verified = TRUE WHERE token = %s', (token,))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def incrementar_intentos_captcha(token, intentos):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE password_reset_tokens SET captcha_attempt = %s WHERE token = %s', (intentos, token))
        conn.commit()
        cursor.close()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()



def obtener_metas(usuario_id):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, nombre, monto_objetivo, monto_actual, creado_en FROM metas WHERE usuario_id = %s', (usuario_id,))
        return cursor.fetchall()
    finally:
        conn.close()

def agregar_meta(usuario_id, nombre, monto_objetivo):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO metas (usuario_id, nombre, monto_objetivo) VALUES (%s, %s, %s)', (usuario_id, nombre, monto_objetivo))
        conn.commit()
        return True
    finally:
        conn.close()

def eliminar_meta_db(id, usuario_id):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM metas WHERE id = %s AND usuario_id = %s', (id, usuario_id))
        conn.commit()
        return True
    finally:
        conn.close()

def actualizar_aporte_meta(meta_id, usuario_id, monto):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE metas SET monto_actual = monto_actual + %s WHERE id = %s AND usuario_id = %s', (monto, meta_id, usuario_id))
        conn.commit()
        return True
    finally:
        conn.close()

def obtener_recurrentes(usuario_id):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, tipo, descripcion, monto, frecuencia, proxima_fecha, activo, creado_en FROM transacciones_recurrentes WHERE usuario_id = %s AND activo = TRUE ORDER BY proxima_fecha ASC, id ASC', (usuario_id,))
        return cursor.fetchall()
    finally:
        conn.close()

def agregar_recurrente(usuario_id, tipo, monto, descripcion, frecuencia, proxima_fecha):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO transacciones_recurrentes (usuario_id, tipo, monto, descripcion, frecuencia, proxima_fecha) 
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (usuario_id, tipo, monto, descripcion, frecuencia, proxima_fecha))
        conn.commit()
        return True
    finally:
        conn.close()

def actualizar_recurrente(id, usuario_id, tipo, monto, descripcion, frecuencia, proxima_fecha):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE transacciones_recurrentes
            SET tipo = %s, monto = %s, descripcion = %s, frecuencia = %s, proxima_fecha = %s
            WHERE id = %s AND usuario_id = %s
        ''', (tipo, monto, descripcion, frecuencia, proxima_fecha, id, usuario_id))
        conn.commit()
        return True
    except Exception:
        return False
    finally:
        conn.close()

def eliminar_recurrente_db(id, usuario_id):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM transacciones_recurrentes WHERE id = %s AND usuario_id = %s', (id, usuario_id))
        conn.commit()
        return True
    finally:
        conn.close()

def actualizar_seguridad_usuario(usuario_id, MFA):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE usuarios SET MFA = %s WHERE id = %s', (MFA, usuario_id))
        conn.commit()
        return True
    finally:
        conn.close()

def procesar_recurrentes_usuario(usuario_id):
    """Verifica si hay transacciones recurrentes que deben ejecutarse hoy o en el pasado."""
    conn = get_connection()
    if conn is None:
        return
    try:
        cursor = conn.cursor(dictionary=True)
        hoy = datetime.now().date()
        cursor.execute('SELECT id, usuario_id, tipo, descripcion, monto, frecuencia, proxima_fecha, activo, creado_en FROM transacciones_recurrentes WHERE usuario_id = %s AND proxima_fecha <= %s AND activo = TRUE', (usuario_id, hoy))
        pendientes = cursor.fetchall()
        
        for p in pendientes:
            # Insertar en la tabla correspondiente
            if p['tipo'] == 'ingreso':
                cursor.execute('INSERT INTO ingresos (usuario_id, monto, descripcion, fecha) VALUES (%s, %s, %s, %s)', 
                             (usuario_id, p['monto'], f"[Recurrente] {p['descripcion']}", p['proxima_fecha']))
            elif p['tipo'] == 'gasto':
                cursor.execute('INSERT INTO gastos (usuario_id, monto, descripcion, fecha) VALUES (%s, %s, %s, %s)', 
                             (usuario_id, p['monto'], f"[Recurrente] {p['descripcion']}", p['proxima_fecha']))
            
            # Calcular proxima fecha
            prox = p['proxima_fecha']
            if p['frecuencia'] == 'mensual':
                # Sumar un mes (aproximado)
                import calendar
                mo = prox.month + 1
                yr = prox.year
                if mo > 12:
                    mo = 1
                    yr += 1
                last_day = calendar.monthrange(yr, mo)[1]
                prox = prox.replace(year=yr, month=mo, day=min(prox.day, last_day))
            elif p['frecuencia'] == 'semanal':
                prox = prox + timedelta(days=7)
            elif p['frecuencia'] == 'diario':
                prox = prox + timedelta(days=1)
            elif p['frecuencia'] == 'anual':
                prox = prox.replace(year=prox.year + 1)
            
            cursor.execute('UPDATE transacciones_recurrentes SET proxima_fecha = %s WHERE id = %s', (prox, p['id']))
            
        conn.commit()
    except Exception as e:
        pass
    finally:
        conn.close()

def actualizar_perfil_usuario_db(usuario_id, nombre_completo='', email='', telefono='', pais='', ciudad='', moneda='USD'):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE usuarios SET nombre_completo = %s, email = %s, telefono = %s, pais = %s, ciudad = %s, moneda = %s WHERE id = %s', 
                       (nombre_completo, email, telefono, pais, ciudad, moneda, usuario_id))
        conn.commit()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def actualizar_foto_perfil_db(usuario_id, foto_bytes):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE usuarios SET foto_perfil = %s WHERE id = %s', (foto_bytes, usuario_id))
        conn.commit()
        return True
    finally:
        conn.close()

def actualizar_nombre_usuario_db(usuario_id, nuevo_nombre):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        # Verificar si ya existe (excluyendo el usuario actual)
        cursor.execute('SELECT id FROM usuarios WHERE nombre_usuario = %s AND id != %s', (nuevo_nombre, usuario_id))
        if cursor.fetchone():
            return False
            
        cursor.execute('UPDATE usuarios SET nombre_usuario = %s WHERE id = %s', (nuevo_nombre, usuario_id))
        conn.commit()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def _redirect_buscar(return_query):
    base = url_for('main.buscar')
    if return_query:
        return redirect(f"{base}?{return_query}")
    return redirect(base)


def _formatear_fecha_ddmmaa(fecha):
    if isinstance(fecha, datetime):
        return fecha.strftime('%d/%m/%Y')
    if isinstance(fecha, date):
        return fecha.strftime('%d/%m/%Y')
    if isinstance(fecha, str):
        try:
            return datetime.fromisoformat(fecha).strftime('%d/%m/%Y')
        except ValueError:
            try:
                return datetime.strptime(fecha, '%Y-%m-%d').strftime('%d/%m/%Y')
            except ValueError:
                return fecha
    return str(fecha)


def calcular_rango_periodo(periodo):
    hoy = date.today()
    if periodo == 'este_mes':
        return hoy.replace(day=1).isoformat(), hoy.isoformat()
    if periodo == 'mes_anterior':
        primero_mes = hoy.replace(day=1)
        ultimo_anterior = primero_mes - timedelta(days=1)
        return ultimo_anterior.replace(day=1).isoformat(), ultimo_anterior.isoformat()
    if periodo == 'ultimos_30':
        return (hoy - timedelta(days=30)).isoformat(), hoy.isoformat()
    if periodo == 'este_anio':
        return hoy.replace(month=1, day=1).isoformat(), date(hoy.year, 12, 31).isoformat()
    return None, None


def calcular_rango_mes_actual():
    hoy = date.today()
    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
    return hoy.replace(day=1).isoformat(), hoy.replace(day=ultimo_dia).isoformat()


def calcular_rango_mes(anio, mes):
    ultimo_dia = calendar.monthrange(int(anio), int(mes))[1]
    return date(int(anio), int(mes), 1).isoformat(), date(int(anio), int(mes), ultimo_dia).isoformat()


def _filtros_desde_request(args):
    filtros = {
        'q': (args.get('q') or '').strip(),
        'desde': args.get('desde') or '',
        'hasta': args.get('hasta') or '',
        'tipo': args.get('tipo') or '',
        'categoria_id': args.get('categoria_id') or '',
        'monto_min': args.get('monto_min') or '',
        'monto_max': args.get('monto_max') or '',
        'orden': args.get('orden') or 'fecha_desc',
        'periodo': args.get('periodo') or '',
    }
    if filtros['periodo'] and not filtros['desde'] and not filtros['hasta']:
        desde, hasta = calcular_rango_periodo(filtros['periodo'])
        if desde and hasta:
            filtros['desde'] = desde
            filtros['hasta'] = hasta
    elif not filtros['desde'] and not filtros['hasta']:
        desde, hasta = calcular_rango_mes_actual()
        filtros['desde'] = desde
        filtros['hasta'] = hasta
    return filtros


def _orden_sql(orden):
    opciones = {
        'fecha_asc': 'fecha ASC, id ASC',
        'monto_desc': 'monto DESC, fecha DESC',
        'monto_asc': 'monto ASC, fecha DESC',
        'descripcion_asc': 'descripcion ASC, fecha DESC',
    }
    return opciones.get(orden, 'fecha DESC, id DESC')


def _sql_union_transacciones(con_categorias=True):
    if con_categorias:
        return """
            SELECT i.id, 'ingreso' AS tipo, i.monto, i.descripcion, i.fecha,
                   i.categoria_id, c.nombre AS categoria_nombre, c.color AS categoria_color
            FROM ingresos i
            LEFT JOIN categorias c ON i.categoria_id = c.id
            WHERE i.usuario_id = %s
            UNION ALL
            SELECT g.id, 'gasto', g.monto, g.descripcion, g.fecha,
                   g.categoria_id, c.nombre, c.color
            FROM gastos g
            LEFT JOIN categorias c ON g.categoria_id = c.id
            WHERE g.usuario_id = %s
            UNION ALL
            SELECT inv.id, 'inversion', inv.monto, inv.descripcion, inv.fecha,
                   inv.categoria_id, c.nombre, c.color
            FROM inversiones inv
            LEFT JOIN categorias c ON inv.categoria_id = c.id
            WHERE inv.usuario_id = %s
        """
    return """
        SELECT id, 'ingreso' AS tipo, monto, descripcion, fecha,
               NULL AS categoria_id, NULL AS categoria_nombre, NULL AS categoria_color
        FROM ingresos WHERE usuario_id = %s
        UNION ALL
        SELECT id, 'gasto', monto, descripcion, fecha,
               NULL, NULL, NULL
        FROM gastos WHERE usuario_id = %s
        UNION ALL
        SELECT id, 'inversion', monto, descripcion, fecha,
               NULL, NULL, NULL
        FROM inversiones WHERE usuario_id = %s
    """


def _where_transacciones(filtros, params):
    condiciones = ['1=1']
    if filtros.get('q'):
        condiciones.append('descripcion LIKE %s')
        params.append(f"%{filtros['q']}%")
    if filtros.get('desde'):
        condiciones.append('fecha >= %s')
        params.append(filtros['desde'])
    if filtros.get('hasta'):
        condiciones.append('fecha <= %s')
        params.append(filtros['hasta'])
    if filtros.get('tipo') in ('ingreso', 'gasto', 'inversion'):
        condiciones.append('tipo = %s')
        params.append(filtros['tipo'])
    if filtros.get('categoria_id'):
        condiciones.append('categoria_id = %s')
        params.append(filtros['categoria_id'])
    if filtros.get('monto_min'):
        try:
            condiciones.append('monto >= %s')
            params.append(float(filtros['monto_min']))
        except ValueError:
            pass
    if filtros.get('monto_max'):
        try:
            condiciones.append('monto <= %s')
            params.append(float(filtros['monto_max']))
        except ValueError:
            pass
    return ' AND '.join(condiciones)


def explorar_transacciones(usuario_id, filtros, pagina=1, por_pagina=25):
    vacio = {
        'resultados': [],
        'resumen': {
            'total_ingresos': 0,
            'total_gastos': 0,
            'total_inversiones': 0,
            'balance': 0,
            'cantidad': 0,
        },
        'paginacion': {'pagina': 1, 'total_paginas': 1, 'total': 0, 'por_pagina': por_pagina},
    }
    conn = get_connection()
    if conn is None:
        return vacio
    try:
        cursor = conn.cursor(dictionary=True)
        base_params = [usuario_id, usuario_id, usuario_id]
        where_params = []
        where_sql = _where_transacciones(filtros, where_params)
        params_base = base_params + where_params
        orden = _orden_sql(filtros.get('orden', 'fecha_desc'))

        for con_categorias in (True, False):
            try:
                union_sql = _sql_union_transacciones(con_categorias)
                subquery = f"SELECT id, tipo, monto, descripcion, fecha, categoria_id, categoria_nombre, categoria_color FROM ({union_sql}) AS t WHERE {where_sql}"

                cursor.execute(
                    f"""
                    SELECT
                        COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) AS total_ingresos,
                        COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) AS total_gastos,
                        COALESCE(SUM(CASE WHEN tipo = 'inversion' THEN monto ELSE 0 END), 0) AS total_inversiones,
                        COUNT(*) AS cantidad
                    FROM ({subquery}) AS resumen
                    """,
                    tuple(params_base),
                )
                resumen_row = cursor.fetchone() or {}
                total_ingresos = float(resumen_row.get('total_ingresos') or 0)
                total_gastos = float(resumen_row.get('total_gastos') or 0)
                total_inversiones = float(resumen_row.get('total_inversiones') or 0)
                cantidad = int(resumen_row.get('cantidad') or 0)
                resumen = {
                    'total_ingresos': total_ingresos,
                    'total_gastos': total_gastos,
                    'total_inversiones': total_inversiones,
                    'balance': total_ingresos - total_gastos - total_inversiones,
                    'cantidad': cantidad,
                }

                total_paginas = max(1, (cantidad + por_pagina - 1) // por_pagina) if cantidad else 1
                pagina_actual = min(max(1, pagina), total_paginas)
                offset = (pagina_actual - 1) * por_pagina

                cursor.execute(
                    f"""
                    SELECT id, tipo, monto, descripcion, fecha, categoria_id,
                           categoria_nombre, categoria_color
                    FROM ({subquery}) AS datos
                    ORDER BY {orden}
                    LIMIT %s OFFSET %s
                    """,
                    tuple(params_base + [por_pagina, offset]),
                )
                resultados = cursor.fetchall()

                return {
                    'resultados': resultados,
                    'resumen': resumen,
                    'paginacion': {
                        'pagina': pagina_actual,
                        'total_paginas': total_paginas,
                        'total': cantidad,
                        'por_pagina': por_pagina,
                    },
                }
            except Exception as e:
                if con_categorias:
                    continue
                print(f"Error explorar_transacciones: {e}")
                return vacio
        return vacio
    finally:
        conn.close()


def obtener_transacciones_filtradas(usuario_id, q, desde, hasta):
    filtros = {'q': q or '', 'desde': desde or '', 'hasta': hasta or '', 'orden': 'fecha_desc'}
    return explorar_transacciones(usuario_id, filtros, pagina=1, por_pagina=10000)['resultados']


def obtener_transacciones_exportacion(usuario_id, filtros):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        where_params = []
        where_sql = _where_transacciones(filtros, where_params)
        orden = _orden_sql(filtros.get('orden', 'fecha_desc'))

        for con_categorias in (True, False):
            try:
                union_sql = _sql_union_transacciones(con_categorias)
                subquery = f"SELECT id, tipo, monto, descripcion, fecha, categoria_id, categoria_nombre, categoria_color FROM ({union_sql}) AS t WHERE {where_sql}"
                cursor.execute(
                    f"""
                    SELECT id, tipo, monto, descripcion, fecha, categoria_id,
                           categoria_nombre, categoria_color
                    FROM ({subquery}) AS datos
                    ORDER BY {orden}
                    """,
                    tuple([usuario_id, usuario_id, usuario_id] + where_params),
                )
                return cursor.fetchall()
            except Exception:
                if con_categorias:
                    continue
                return []
        return []
    finally:
        conn.close()

def obtener_categorias(usuario_id):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id, usuario_id, nombre, tipo, color, creado_en FROM categorias WHERE usuario_id = %s ORDER BY tipo ASC, nombre ASC', (usuario_id,))
        return cursor.fetchall()
    finally:
        conn.close()

def validar_categoria_para_tipo(usuario_id, categoria_id, tipo_transaccion):
    """
    Valida que una categoría sea válida para el tipo de transacción especificado.
    
    Args:
        usuario_id: ID del usuario
        categoria_id: ID de la categoría (puede ser None)
        tipo_transaccion: Tipo de transacción ('ingreso', 'gasto', 'inversion')
    
    Returns:
        tuple: (es_valida, mensaje_error)
    """
    # Si no hay categoría, es válido (las categorías son opcionales)
    if not categoria_id:
        return True, ""
    
    # Validar que el tipo de transacción sea válido
    tipos_validos = ['ingreso', 'gasto', 'inversion']
    if tipo_transaccion not in tipos_validos:
        return False, f"Tipo de transacción inválido: {tipo_transaccion}"
    
    conn = get_connection()
    if conn is None:
        return False, "Error de conexión a la base de datos"
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Obtener la categoría
        cursor.execute(
            'SELECT id, usuario_id, nombre, tipo, color, creado_en FROM categorias WHERE id = %s AND usuario_id = %s',
            (categoria_id, usuario_id)
        )
        categoria = cursor.fetchone()
        
        if not categoria:
            return False, "Categoría no encontrada"
        
        # Validar que el tipo de la categoría coincida con el tipo de transacción
        tipo_categoria = categoria.get('tipo', '').lower()
        
        if tipo_categoria and tipo_categoria != tipo_transaccion:
            return False, f"La categoría '{categoria['nombre']}' es para {tipo_categoria}s, no para {tipo_transaccion}s"
        
        return True, ""
    finally:
        conn.close()

def agregar_categoria(usuario_id, nombre, tipo, color):
    conn = get_connection()
    if conn is None:
        return False
    try:
        # Validar que el tipo sea válido
        tipos_validos = ['ingreso', 'gasto', 'inversion']
        if tipo not in tipos_validos:
            return False
        
        cursor = conn.cursor()
        cursor.execute('INSERT INTO categorias (usuario_id, nombre, tipo, color) VALUES (%s, %s, %s, %s)', 
                      (usuario_id, nombre, tipo, color))
        conn.commit()
        return True
    finally:
        conn.close()

def eliminar_categoria_db(id, usuario_id):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM categorias WHERE id = %s AND usuario_id = %s', (id, usuario_id))
        conn.commit()
        return True
    finally:
        conn.close()

def obtener_presupuestos_detallados(usuario_id):
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        # Obtener presupuestos con el nombre de la categoría y el gasto actual del mes
        hoy = datetime.now()
        desde_mes, hasta_mes = calcular_rango_mes(hoy.year, hoy.month)
        
        cursor.execute('''
            SELECT p.*, c.nombre as categoria_nombre, 
            (SELECT COALESCE(SUM(monto), 0) FROM gastos 
             WHERE usuario_id = %s AND categoria_id = p.categoria_id 
             AND fecha BETWEEN %s AND %s) as gasto_actual
            FROM presupuestos p
            JOIN categorias c ON p.categoria_id = c.id
            WHERE p.usuario_id = %s
        ''', (usuario_id, desde_mes, hasta_mes, usuario_id))
        
        presupuestos = cursor.fetchall()
        for p in presupuestos:
            p['porcentaje'] = round((float(p['gasto_actual']) / float(p['monto']) * 100), 1) if p['monto'] > 0 else 0
        return presupuestos
    finally:
        conn.close()

def obtener_presupuestos_simples(usuario_id):
    """Obtiene presupuestos con gastos reales del mes actual"""
    conn = get_connection()
    if conn is None:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            'SELECT p.id, p.usuario_id, p.categoria_id, p.monto, p.gastado, c.nombre AS categoria_nombre '
            'FROM presupuestos p '
            'LEFT JOIN categorias c ON c.id = p.categoria_id '
            'WHERE p.usuario_id = %s',
            (usuario_id,)
        )
        presupuestos = cursor.fetchall()

        # Obtener gastos totales del mes actual una sola vez
        desde_mes, hasta_mes = calcular_rango_mes(datetime.now().year, datetime.now().month)
        cursor.execute(
            'SELECT COALESCE(SUM(monto), 0) AS total_gastos '
            'FROM gastos '
            'WHERE usuario_id = %s AND fecha BETWEEN %s AND %s',
            (usuario_id, desde_mes, hasta_mes)
        )
        gasto_total_result = cursor.fetchone() or {}
        gasto_total = float(gasto_total_result.get('total_gastos') or 0)
        
        for p in presupuestos:
            p['categoria_nombre'] = p.get('categoria_nombre') or 'Sin categoría'
            p['monto'] = float(p.get('monto') or 0)
            p['gastado'] = float(p.get('gastado') or 0)
            
            # Por simplicidad, distribuir los gastos totales entre los presupuestos
            # En un futuro, esto debería conectarse con categorías reales
            if len(presupuestos) > 0:
                p['gasto_actual'] = gasto_total / len(presupuestos)
            else:
                p['gasto_actual'] = 0
            
            # Calcular porcentaje usando gastado si está disponible, sino usar gasto_actual
            gasto_para_calculo = p['gastado'] if p['gastado'] > 0 else p['gasto_actual']
            p['porcentaje'] = round((gasto_para_calculo / p['monto'] * 100), 1) if p['monto'] > 0 else 0
            # Para la progress bar, limitar a 100 visualmente pero guardar el valor real
            p['porcentaje_visual'] = min(p['porcentaje'], 100)
        
        return presupuestos
    finally:
        conn.close()

def agregar_o_actualizar_presupuesto(usuario_id, categoria_id, monto):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT id FROM presupuestos WHERE usuario_id = %s AND categoria_id = %s', (usuario_id, categoria_id))
        existente = cursor.fetchone()
        
        if existente:
            cursor.execute('UPDATE presupuestos SET monto = %s WHERE id = %s', (monto, existente['id']))
        else:
            cursor.execute('INSERT INTO presupuestos (usuario_id, categoria_id, monto) VALUES (%s, %s, %s)', 
                          (usuario_id, categoria_id, monto))
        conn.commit()
        return True
    finally:
        conn.close()

def actualizar_totp_secret_db(usuario_id, secret):
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE usuarios SET totp_secret = %s WHERE id = %s', (secret, usuario_id))
        conn.commit()
        return True
    finally:
        conn.close()
def actualizar_ultimo_login(usuario_id):
    """Actualiza la fecha y hora del último login del usuario"""
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE usuarios SET ultimo_login = NOW() WHERE id = %s', (usuario_id,))
        conn.commit()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def actualizar_presupuesto(usuario_id, presupuesto_id, monto):
    """Actualiza el monto de un presupuesto"""
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE presupuestos SET monto = %s WHERE id = %s AND usuario_id = %s', 
                      (monto, presupuesto_id, usuario_id))
        conn.commit()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def actualizar_presupuesto_gastado(usuario_id, presupuesto_id, gastado):
    """Actualiza el gastado de un presupuesto"""
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('UPDATE presupuestos SET gastado = %s WHERE id = %s AND usuario_id = %s', 
                      (gastado, presupuesto_id, usuario_id))
        conn.commit()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

def eliminar_presupuesto_db(usuario_id, presupuesto_id):
    """Elimina un presupuesto"""
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM presupuestos WHERE id = %s AND usuario_id = %s', 
                      (presupuesto_id, usuario_id))
        conn.commit()
        return True
    except Exception as e:
        return False
    finally:
        conn.close()

# --- Funciones auxiliares para MFA/TOTP ---
def validar_totp(totp_secret, codigo):
    """Valida un código TOTP contra el secreto del usuario"""
    try:
        totp = pyotp.TOTP(totp_secret)
        # Permitir ventana de ±30 segundos (1 ventana antes y después)
        return totp.verify(codigo, valid_window=1)
    except Exception as e:
        return False

def obtener_totp_secret_usuario(usuario_id):
    """Obtiene el secreto TOTP de un usuario"""
    conn = get_connection()
    if conn is None:
        return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT totp_secret FROM usuarios WHERE id = %s', (usuario_id,))
        resultado = cursor.fetchone()
        cursor.close()
        return resultado['totp_secret'] if resultado else None
    except Exception as e:
        return None
    finally:
        conn.close()

def usuario_tiene_mfa_activado(usuario_id):
    """Verifica si un usuario tiene MFA activado"""
    conn = get_connection()
    if conn is None:
        return False
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT MFA FROM usuarios WHERE id = %s', (usuario_id,))
        resultado = cursor.fetchone()
        cursor.close()
        return resultado['MFA'] if resultado else False
    except Exception as e:
        return False
    finally:
        conn.close()